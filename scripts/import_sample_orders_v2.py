#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import multi-channel sample order files into control_tower.sample_orders.

v2.8.1 fix:
- Do NOT mark "Đang trung chuyển" as cancelled.
- Cancellation detection uses explicit cancelled statuses only, not substring matching.

Supported minimal columns:
- Mã đơn hàng / external_order_id
- Ngày tạo đơn / order_date
- Tình trạng / order_status
- Giá vốn / cogs_amount

Usage:
  python scripts/import_sample_orders_v2.py --file "data/sample_orders/2026-06/TIKTOK/sample.xlsx" --channel TIKTOK --rerun
"""

from __future__ import annotations

import argparse
import json
import os
import re
import uuid
from datetime import datetime, date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

try:
    import psycopg2
    from psycopg2.extras import Json
except ImportError as exc:
    raise SystemExit("Missing psycopg2. Install with: pip install psycopg2-binary") from exc


HEADER_ALIASES = {
    "external_order_id": [
        "mã đơn hàng", "ma don hang", "mã đơn", "ma don", "order id", "order_id",
        "external_order_id", "mã đơn hàng tiktok", "mã đơn hàng shopee"
    ],
    "order_date": [
        "ngày tạo đơn", "ngay tao don", "ngày tạo", "ngay tao", "order date",
        "order_date", "ngày", "date"
    ],
    "order_status": [
        "tình trạng", "tinh trang", "trạng thái", "trang thai", "order status",
        "order_status", "status"
    ],
    "cogs_amount": [
        "giá vốn", "gia von", "cogs", "cogs_amount", "giá vốn mẫu",
        "gia von mau", "cost", "cost amount"
    ],
    "sample_qty": [
        "số lượng", "so luong", "quantity", "qty", "total_sku_qty", "sl"
    ],
    "sku_code": [
        "sku", "mã sku", "ma sku", "mã hàng", "ma hang", "seller sku", "seller_sku"
    ],
    "product_name": [
        "tên sản phẩm", "ten san pham", "product name", "product_name", "tên hàng", "ten hang"
    ],
    "shop_label": [
        "shop", "shop label", "shop_label", "gian hàng", "gian hang"
    ],
    "source_channel": [
        "kênh", "kenh", "channel", "source_channel", "sàn", "san"
    ],
}

REQUIRED = ["external_order_id", "order_date", "cogs_amount"]


def norm_text(s: Any) -> str:
    if s is None:
        return ""
    text = str(s).strip().lower()
    repl = {
        "đ": "d",
        "á": "a", "à": "a", "ả": "a", "ã": "a", "ạ": "a",
        "ă": "a", "ắ": "a", "ằ": "a", "ẳ": "a", "ẵ": "a", "ặ": "a",
        "â": "a", "ấ": "a", "ầ": "a", "ẩ": "a", "ẫ": "a", "ậ": "a",
        "é": "e", "è": "e", "ẻ": "e", "ẽ": "e", "ẹ": "e",
        "ê": "e", "ế": "e", "ề": "e", "ể": "e", "ễ": "e", "ệ": "e",
        "í": "i", "ì": "i", "ỉ": "i", "ĩ": "i", "ị": "i",
        "ó": "o", "ò": "o", "ỏ": "o", "õ": "o", "ọ": "o",
        "ô": "o", "ố": "o", "ồ": "o", "ổ": "o", "ỗ": "o", "ộ": "o",
        "ơ": "o", "ớ": "o", "ờ": "o", "ở": "o", "ỡ": "o", "ợ": "o",
        "ú": "u", "ù": "u", "ủ": "u", "ũ": "u", "ụ": "u",
        "ư": "u", "ứ": "u", "ừ": "u", "ử": "u", "ữ": "u", "ự": "u",
        "ý": "y", "ỳ": "y", "ỷ": "y", "ỹ": "y", "ỵ": "y",
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    text = re.sub(r"\s+", " ", text)
    return text


def build_alias_lookup() -> Dict[str, str]:
    out = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for a in aliases:
            out[norm_text(a)] = canonical
    return out


ALIAS_LOOKUP = build_alias_lookup()


def parse_date(value: Any) -> date:
    if value is None or value == "":
        raise ValueError("missing date")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Cannot parse date: {value!r}")


def to_num(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    text = str(value).strip().replace(",", "")
    text = re.sub(r"[^\d\.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return default
    return float(text)


def to_bool_cancelled(status: Any) -> bool:
    """Return True only for explicit cancelled/returned statuses.

    Important: do not use substring matching on 'huy' because 'Đang trung chuyển'
    normalizes to 'dang trung chuyen' and contains the letters h-u-y.
    """
    text = norm_text(status)
    if not text:
        return False

    cancelled_exact = {
        "huy", "da huy", "don huy", "huy don", "cancelled", "canceled", "cancel", "da huy don"
    }
    returned_exact = {
        "hoan", "da hoan", "tra hang", "hoan hang", "returned", "return"
    }
    if text in cancelled_exact or text in returned_exact:
        return True

    # Phrases with clear word boundaries.
    patterns = [
        r"(^|\s)da huy($|\s)",
        r"(^|\s)huy don($|\s)",
        r"(^|\s)don huy($|\s)",
        r"(^|\s)cancelled($|\s)",
        r"(^|\s)canceled($|\s)",
        r"(^|\s)returned($|\s)",
        r"(^|\s)da hoan($|\s)",
    ]
    return any(re.search(p, text) for p in patterns)


def normalize_channel(channel: str) -> str:
    text = norm_text(channel).upper()
    if text in {"DIGITAL", "NHANH"}:
        return "NHANH"
    if text in {"SHOPEE", "SHOPMALL"}:
        return "SHOPEE"
    if text in {"TIKTOK", "TIKTOKSHOP", "TIKTOK SHOP"}:
        return "TIKTOK"
    return text or "UNKNOWN"


def find_header(ws) -> tuple[int, Dict[str, int]]:
    best_score = -1
    best_row = None
    best_map: Dict[str, int] = {}
    for r in range(1, min(ws.max_row, 20) + 1):
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        header_map: Dict[str, int] = {}
        for idx, val in enumerate(row_values):
            key = ALIAS_LOOKUP.get(norm_text(val))
            if key and key not in header_map:
                header_map[key] = idx
        score = sum(1 for req in REQUIRED if req in header_map)
        if score > best_score:
            best_score = score
            best_row = r
            best_map = header_map
        if score == len(REQUIRED):
            return r, header_map
    raise RuntimeError(
        f"Could not find required sample order headers. Required: {REQUIRED}. "
        f"Best row={best_row}, found={list(best_map.keys())}."
    )


def read_sample_file(path: Path, channel: str, shop_label: Optional[str], sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row, header_map = find_header(ws)
    rows: List[Dict[str, Any]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        def get(field, default=None):
            idx = header_map.get(field)
            return values[idx] if idx is not None and idx < len(values) else default

        external_order_id = str(get("external_order_id", "") or "").strip()
        if not external_order_id:
            continue

        order_date = parse_date(get("order_date"))
        order_status = str(get("order_status", "") or "").strip()
        file_channel = str(get("source_channel", "") or "").strip()
        file_shop = str(get("shop_label", "") or "").strip()
        sku_code = str(get("sku_code", "") or "").strip()
        product_name = str(get("product_name", "") or "").strip()

        source_channel = normalize_channel(file_channel or channel)
        normalized_channel = normalize_channel(source_channel)
        final_shop_label = shop_label or file_shop or source_channel

        sample_qty = to_num(get("sample_qty"), default=1.0)
        if sample_qty == 0:
            sample_qty = 1.0

        cogs_amount = to_num(get("cogs_amount"), default=0.0)
        sample_sku_count = 1 if not sku_code else 1
        is_cancelled = to_bool_cancelled(order_status)

        raw_payload = {
            "external_order_id": external_order_id,
            "order_date": order_date.isoformat(),
            "order_status": order_status,
            "source_channel": source_channel,
            "shop_label": final_shop_label,
            "sku_code": sku_code,
            "product_name": product_name,
            "sample_qty": sample_qty,
            "sample_sku_count": sample_sku_count,
            "sample_cogs_amount": cogs_amount,
            "is_cancelled": is_cancelled,
            "source_row_no": r,
        }

        rows.append({
            "external_order_id": external_order_id,
            "shop_label": final_shop_label,
            "order_date": order_date,
            "order_status": order_status,
            "is_cancelled": is_cancelled,
            "brand_group": "SAMPLE",
            "total_sku_qty": sample_qty,
            "cogs_amount": cogs_amount,
            "source_channel": source_channel,
            "normalized_channel": normalized_channel,
            "sample_qty": sample_qty,
            "sample_sku_count": sample_sku_count,
            "sample_cogs_amount": cogs_amount,
            "import_month": order_date.replace(day=1),
            "source_file": str(path),
            "source_row_no": r,
            "raw_payload": raw_payload,
        })

    return rows


def connect():
    return psycopg2.connect(
        host=os.getenv("PGHOST", "localhost"),
        port=int(os.getenv("PGPORT", "5433")),
        dbname=os.getenv("PGDATABASE", "DuLieu"),
        user=os.getenv("PGUSER", "postgres"),
        password=os.getenv("PGPASSWORD"),
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--channel", required=True, help="TIKTOK, SHOPEE, NHANH/DIGITAL")
    ap.add_argument("--shop-label", default=None)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--rerun", action="store_true")
    args = ap.parse_args()

    path = Path(args.file)
    rows = read_sample_file(path, args.channel, args.shop_label, args.sheet)
    if not rows:
        raise SystemExit("No sample order rows found.")

    load_batch_id = str(uuid.uuid4())
    dates = sorted({r["order_date"] for r in rows})
    normalized_channel = normalize_channel(args.channel)
    shop_label = args.shop_label or rows[0]["shop_label"]

    insert_sql = """
    INSERT INTO control_tower.sample_orders (
        external_order_id, shop_label, order_date, order_status, is_cancelled, brand_group,
        total_sku_qty, cogs_amount, source_channel, normalized_channel, sample_qty,
        sample_sku_count, sample_cogs_amount, import_month, source_file, source_row_no,
        load_batch_id, raw_payload, updated_at
    ) VALUES (
        %(external_order_id)s, %(shop_label)s, %(order_date)s, %(order_status)s, %(is_cancelled)s, %(brand_group)s,
        %(total_sku_qty)s, %(cogs_amount)s, %(source_channel)s, %(normalized_channel)s, %(sample_qty)s,
        %(sample_sku_count)s, %(sample_cogs_amount)s, %(import_month)s, %(source_file)s, %(source_row_no)s,
        %(load_batch_id)s, %(raw_payload)s, now()
    )
    """

    conn = connect()
    try:
        with conn:
            with conn.cursor() as cur:
                if args.rerun:
                    cur.execute(
                        """
                        DELETE FROM control_tower.sample_orders
                        WHERE order_date = ANY(%s)
                          AND normalized_channel = %s
                          AND COALESCE(shop_label, '') = COALESCE(%s, '')
                        """,
                        (dates, normalized_channel, shop_label),
                    )

                for r in rows:
                    data = dict(r)
                    data["load_batch_id"] = load_batch_id
                    data["raw_payload"] = Json(data["raw_payload"])
                    cur.execute(insert_sql, data)
    finally:
        conn.close()

    active_rows = [r for r in rows if not r["is_cancelled"]]
    cancelled_rows = [r for r in rows if r["is_cancelled"]]
    summary = {
        "status": "success",
        "file": str(path),
        "channel": normalized_channel,
        "shop_label": shop_label,
        "rows_loaded": len(rows),
        "active_rows": len(active_rows),
        "cancelled_rows": len(cancelled_rows),
        "dates": [d.isoformat() for d in dates],
        "sample_qty": sum(float(r["sample_qty"]) for r in active_rows),
        "sample_cogs_amount": sum(float(r["sample_cogs_amount"]) for r in active_rows),
        "load_batch_id": load_batch_id,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
